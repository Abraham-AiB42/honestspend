"""Parse issuer activity workbooks (Amex activity.xlsx and siblings).

These are not the legacy Budget.xlsx daily ledger. Banks download them as
activity.xlsx / activity (1).xlsx — product name and account mask live in
the first rows, transactions start at a Date/Description/Amount header.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

from sqlalchemy.orm import Session

from honestspend.db import Account, Transaction
from honestspend.services.merchant_catalog import format_import_memo


def _as_date(val: Any) -> date | None:
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y", "%m-%d-%Y"):
        try:
            return datetime.strptime(s[:10] if fmt == "%Y-%m-%d" else s, fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(s[:10]).date()
    except ValueError:
        return None


def _as_decimal(val: Any) -> Decimal | None:
    if val is None or val == "":
        return None
    if isinstance(val, Decimal):
        return val
    if isinstance(val, (int, float)):
        return Decimal(str(val))
    s = str(val).strip().replace("$", "").replace(",", "")
    if not s:
        return None
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def _cell_str(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()


def looks_like_activity_xlsx(rows: list[tuple[Any, ...]] | None = None, sheet_names: list[str] | None = None) -> bool:
    names = " ".join(sheet_names or []).lower()
    if "transaction details" in names or "transaction summary" in names:
        return True
    blob = " ".join(_cell_str(c) for r in (rows or [])[:12] for c in r[:4]).lower()
    if "account number" in blob and re.search(r"x{2,}", blob):
        return True
    if re.search(r"\b(blue business|hilton honors|amazon business|gold card|platinum card)\b", blob):
        return True
    return False


def _norm_header(h: str) -> str:
    h = _cell_str(h).strip().lower().replace(".", "")
    return re.sub(r"\s+", " ", h)


def _header_indexes(row: tuple[Any, ...]) -> dict[str, int] | None:
    norms = [_norm_header(c) for c in row]
    idx: dict[str, int] = {}
    for i, h in enumerate(norms):
        if h in ("date", "trans date", "transaction date", "posted date", "post date") and "date" not in idx:
            # Prefer trans/transaction date over post date
            if h == "post date" and "date" not in idx:
                idx["date"] = i
            elif h != "post date":
                idx["date"] = i
        elif h in ("description", "payee", "name", "merchant") and "desc" not in idx:
            idx["desc"] = i
        elif h in ("amount", "amt", "debit") and "amount" not in idx:
            idx["amount"] = i
        elif h == "category" and "category" not in idx:
            idx["category"] = i
        elif h in ("reference", "ref", "trans id") and "reference" not in idx:
            idx["reference"] = i
    if "date" in idx and "desc" in idx and "amount" in idx:
        return idx
    return None


def _load_openpyxl_sheets(content: bytes) -> list[tuple[str, list[tuple[Any, ...]]]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        out: list[tuple[str, list[tuple[Any, ...]]]] = []
        for name in wb.sheetnames:
            rows: list[tuple[Any, ...]] = []
            for row in wb[name].iter_rows(values_only=True):
                rows.append(tuple(row))
                if len(rows) > 8000:
                    break
            out.append((name, rows))
        return out
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _load_xlrd_sheets(content: bytes) -> list[tuple[str, list[tuple[Any, ...]]]]:
    import xlrd

    book = xlrd.open_workbook(file_contents=content)
    out: list[tuple[str, list[tuple[Any, ...]]]] = []
    for name in book.sheet_names():
        sh = book.sheet_by_name(name)
        rows: list[tuple[Any, ...]] = []
        for r in range(min(sh.nrows, 8000)):
            cells: list[Any] = []
            for c in range(sh.ncols):
                cell = sh.cell(r, c)
                val: Any = cell.value
                if cell.ctype == getattr(xlrd, "XL_CELL_DATE", 3):
                    try:
                        val = xlrd.xldate_as_datetime(cell.value, book.datemode).date()
                    except Exception:
                        pass
                cells.append(val)
            rows.append(tuple(cells))
        out.append((name, rows))
    return out


def _load_html_tables(content: bytes) -> list[tuple[str, list[tuple[Any, ...]]]]:
    from html.parser import HTMLParser

    class _Tables(HTMLParser):
        def __init__(self) -> None:
            super().__init__(convert_charrefs=True)
            self.tables: list[list[list[str]]] = []
            self._table: list[list[str]] | None = None
            self._row: list[str] | None = None
            self._cell: list[str] | None = None

        def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
            t = tag.lower()
            if t == "table":
                self._table = []
            elif t == "tr" and self._table is not None:
                self._row = []
            elif t in ("td", "th") and self._row is not None:
                self._cell = []

        def handle_endtag(self, tag: str) -> None:
            t = tag.lower()
            if t in ("td", "th") and self._cell is not None and self._row is not None:
                self._row.append("".join(self._cell).strip())
                self._cell = None
            elif t == "tr" and self._row is not None and self._table is not None:
                self._table.append(self._row)
                self._row = None
            elif t == "table" and self._table is not None:
                if any(any(c for c in row) for row in self._table):
                    self.tables.append(self._table)
                self._table = None

        def handle_data(self, data: str) -> None:
            if self._cell is not None:
                self._cell.append(data)

    raw = content
    if raw[:2] == b"\xff\xfe":
        text = raw.decode("utf-16-le", errors="ignore")
    elif raw[:2] == b"\xfe\xff":
        text = raw.decode("utf-16-be", errors="ignore")
    else:
        text = raw.decode("utf-8", errors="ignore")
        if "\x00" in text[:200]:
            text = raw.decode("utf-16-le", errors="ignore")
    if "<table" not in text.lower() and "<td" not in text.lower():
        return []
    parser = _Tables()
    try:
        parser.feed(text)
        parser.close()
    except Exception:
        return []
    out: list[tuple[str, list[tuple[Any, ...]]]] = []
    for i, table in enumerate(parser.tables):
        rows = [tuple(r) for r in table[:8000]]
        blob = " ".join(_cell_str(c) for r in rows[:8] for c in r).lower()
        if any(_header_indexes(r) for r in rows[:12]):
            name = "Transaction Details"
        elif "account ending" in blob or "account number" in blob:
            name = "Account Header"
        elif "summary" in blob:
            name = "Transaction Summary"
        else:
            name = f"Sheet{i + 1}"
        out.append((name, rows))
    return out


def load_workbook_sheets(content: bytes) -> list[tuple[str, list[tuple[Any, ...]]]]:
    """Load .xlsx, real .xls (BIFF), or HTML-saved-as-xls into named row lists."""
    if not content:
        return []
    # ZIP / xlsx
    if content[:2] == b"PK":
        try:
            return _load_openpyxl_sheets(content)
        except Exception:
            pass
    # OLE compound document (Excel 97–2003)
    if content[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        try:
            return _load_xlrd_sheets(content)
        except Exception:
            pass
    html = _load_html_tables(content)
    if html:
        return html
    try:
        return _load_openpyxl_sheets(content)
    except Exception:
        pass
    try:
        return _load_xlrd_sheets(content)
    except Exception:
        return []


def parse_activity_xlsx(content: bytes, filename: str = "activity.xlsx") -> dict[str, Any] | None:
    """Return one account + rows, or None if this is not an issuer activity book."""
    sheets = load_workbook_sheets(content)
    if not sheets:
        return None

    try:
        sheet_names = [n for n, _ in sheets]
        details_rows: list[tuple[Any, ...]] | None = None
        summary_rows: list[tuple[Any, ...]] | None = None
        header_i = None
        cols = None
        best_count = -1
        for name, rows in sheets:
            low = name.lower()
            for i, row in enumerate(rows[:40]):
                found = _header_indexes(row)
                if not found:
                    continue
                ntx = 0
                for later in rows[i + 1 :]:
                    if found["date"] < len(later) and _as_date(later[found["date"]]):
                        ntx += 1
                if ntx > best_count:
                    best_count = ntx
                    details_rows = rows
                    header_i = i
                    cols = found
            if "summary" in low:
                summary_rows = rows
        if details_rows is None:
            details_rows = sheets[0][1]

        raw_rows = details_rows[:8000]
        all_blob = " ".join(
            _cell_str(c) for _, rows in sheets for r in rows[:20] for c in r
        )

        if not looks_like_activity_xlsx(raw_rows, sheet_names):
            if header_i is None and "account ending" not in all_blob.lower():
                return None

        product = ""
        cardholder = ""
        acct_mask = ""
        ending = None
        for _, rows in sheets:
            for i, row in enumerate(rows[:40]):
                cells = [_cell_str(c) for c in row]
                if i == 0 and len(cells) > 1 and cells[1] and "card" in cells[1].lower():
                    product = product or cells[1].split("/")[0].strip()
                if cells and cells[0].lower() == "prepared for" and i + 1 < len(rows):
                    cardholder = cardholder or _cell_str(rows[i + 1][0] if rows[i + 1] else "")
                if cells and cells[0].lower() == "account number" and i + 1 < len(rows):
                    acct_mask = acct_mask or _cell_str(rows[i + 1][0] if rows[i + 1] else "")
                if cells and "account ending" in cells[0].lower():
                    acct_mask = acct_mask or re.sub(r"\s+", " ", cells[0])
                    if len(cells) > 1 and cells[1]:
                        acct_mask = acct_mask + " " + cells[1]
                if not product:
                    for c in cells[1:3]:
                        if c and len(c) > 8 and "card" in c.lower():
                            product = c.split("/")[0].strip()
                            break
                if cells and cells[0].lower() in ("total balance", "new balance"):
                    for c in row[1:]:
                        d = _as_decimal(c)
                        if d is not None:
                            ending = d
                            break

        last4 = None
        acct_tail = None
        m = re.search(r"(\d{4,5})\s*$", (acct_mask or "").replace(" ", "").replace("\n", ""))
        if m:
            last4 = m.group(1)[-4:]
            acct_tail = m.group(1)
        if not last4:
            m2 = re.search(r"ending\s+in\s+(\d{4})", all_blob, re.I)
            if m2:
                last4 = m2.group(1)
                acct_tail = last4
                acct_mask = acct_mask or f"ending in {last4}"

        rows: list[dict[str, Any]] = []
        categories: list[str] = []
        if header_i is not None and cols is not None:
            for row in raw_rows[header_i + 1 :]:
                if cols["date"] >= len(row):
                    continue
                d = _as_date(row[cols["date"]])
                if not d:
                    continue
                payee = _cell_str(row[cols["desc"]]) if cols["desc"] < len(row) else ""
                amt = _as_decimal(row[cols["amount"]]) if cols["amount"] < len(row) else None
                if amt is None:
                    continue
                cat = ""
                if "category" in cols and cols["category"] < len(row):
                    cat = _cell_str(row[cols["category"]])
                    if cat:
                        categories.append(cat)
                ref = ""
                if "reference" in cols and cols["reference"] < len(row):
                    ref = _cell_str(row[cols["reference"]])
                rows.append(
                    {
                        "txn_date": d,
                        "payee": payee,
                        "amount": amt,
                        "category": cat,
                        "bank_txn_id": ref,
                    }
                )

        dates = [r["txn_date"] for r in rows]
        prod_l = product.lower()
        is_biz = bool(
            re.search(
                r"\bbusiness\b|\bamazon business\b|\bblue business\b",
                prod_l,
            )
        )
        ext = (filename.rsplit(".", 1)[-1] if "." in filename else "xlsx").lower()
        org = "American Express"
        fn = (filename or "").lower()
        blob_l = all_blob.lower()
        if fn.startswith("dfs") or "discover" in fn or "directpay" in blob_l:
            org = "Discover"
        elif "chase" in fn:
            org = "Chase"
        return {
            "ok": True,
            "file_format": "xls" if ext == "xls" else "xlsx",
            "kind": "credit",
            "org": org,
            "product": product or None,
            "cardholder": cardholder or None,
            "acct_mask": acct_mask or None,
            "acctid": acct_tail or last4,
            "last4": last4,
            "suggested_entity_type": "business" if is_biz else "personal",
            "transactions_found": len(rows),
            "ledger_balance": str(ending) if ending is not None else None,
            "date_from": min(dates).isoformat() if dates else None,
            "date_to": max(dates).isoformat() if dates else None,
            "sample_payees": [r["payee"] for r in rows[:8] if r.get("payee")],
            "sample_categories": categories[:12],
            "rows": rows,
            "filename": filename,
        }
    except Exception:
        return None


@dataclass
class ActivityXlsxImportResult:
    rows_scanned: int = 0
    transactions_created: int = 0
    skipped_existing: int = 0
    skipped_bad: int = 0
    errors: list[str] = field(default_factory=list)
    ending_balance: str | None = None


def import_activity_xlsx(
    session: Session,
    *,
    account_id: int,
    content: bytes,
    filename: str,
) -> ActivityXlsxImportResult:
    """Write parsed issuer activity rows onto an existing account."""
    result = ActivityXlsxImportResult()
    parsed = parse_activity_xlsx(content, filename)
    if not parsed:
        result.errors.append("Not an issuer activity workbook")
        return result
    acct = session.get(Account, account_id)
    if acct is None:
        result.errors.append("Account not found")
        return result

    from honestspend.services.account_balance import apply_amount_to_account
    from honestspend.services.import_amounts import normalize_credit_import_amount
    from honestspend.services.import_dedupe import load_dedupe_index
    from honestspend.services.import_txn_review import find_existing_by_external_id

    dedupe = load_dedupe_index(session, account_id)
    for row in parsed.get("rows") or []:
        result.rows_scanned += 1
        d = row.get("txn_date")
        payee = (row.get("payee") or "").strip()
        amt = row.get("amount")
        if not d or amt is None:
            result.skipped_bad += 1
            continue
        amount = amt if isinstance(amt, Decimal) else Decimal(str(amt))
        if acct.kind == "credit":
            amount = normalize_credit_import_amount(amount, payee)
        bank_txn_id = (row.get("bank_txn_id") or "").strip()
        external_id = dedupe.preferred_external_id(
            txn_date=d,
            amount=amount,
            payee=payee,
            bank_txn_id=bank_txn_id or None,
            source="xlsx",
        )
        existing = find_existing_by_external_id(session, account_id, external_id)
        if existing is not None:
            result.skipped_existing += 1
            continue
        if dedupe.is_duplicate(
            txn_date=d,
            amount=amount,
            payee=payee,
            bank_txn_id=bank_txn_id or None,
        ):
            result.skipped_existing += 1
            continue
        session.add(
            Transaction(
                profile_id=acct.profile_id,
                account_id=acct.id,
                txn_date=d,
                amount=amount,
                payee=payee or None,
                memo=format_import_memo(
                    "XLSX",
                    filename,
                    category=str(row.get("category") or ""),
                ),
                status="cleared",
                external_id=external_id,
            )
        )
        apply_amount_to_account(acct, amount)
        dedupe.remember(
            txn_date=d,
            amount=amount,
            payee=payee,
            bank_txn_id=bank_txn_id or None,
            external_id=external_id,
        )
        result.transactions_created += 1

    if parsed.get("ledger_balance"):
        try:
            from honestspend.services.reconcile import set_institution_balance

            bal = Decimal(str(parsed["ledger_balance"]))
            set_institution_balance(session, account_id, bal, mark_reconciled=False)
            result.ending_balance = str(bal)
        except Exception:
            pass
    session.flush()
    try:
        from honestspend.services.import_reminders import mark_import_activity

        mark_import_activity(session)
    except Exception:
        pass
    return result
