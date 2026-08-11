"""Import historical Budget.xlsx daily ledger into Financial OS."""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from financial_os.db import Account, Category, Profile, Transaction

# Legacy Excel column headers (row 2) → personal category codes
LEGACY_COLUMN_MAP: dict[str, str] = {
    "food": "PER_GROCERIES",
    "purch": "PER_SHOPPING",
    "gas": "PER_GAS",
    "ins": "PER_INSURANCE",
    "home": "PER_HOUSING",
    "util": "PER_UTILITIES",
    "net": "PER_UTILITIES",  # internet often under util
    "cell": "PER_CELL",
    "income": "PER_INCOME_WAGES",
    "save": "SYS_TRANSFER",
    "daily": "",  # daily total — skip, derived
}


@dataclass
class ImportResult:
    rows_scanned: int = 0
    transactions_created: int = 0
    skipped_empty: int = 0
    skipped_existing: int = 0
    errors: list[str] = field(default_factory=list)
    date_from: date | None = None
    date_to: date | None = None
    schedules_advanced: int = 0
    schedules_advanced_names: list[str] = field(default_factory=list)
    schedule_advance_hint: str | None = None
    schedule_advance_error: str | None = None


def _norm_header(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip().lower()


def _as_date(val: Any) -> date | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        try:
            return datetime.fromisoformat(val[:10]).date()
        except ValueError:
            return None
    return None


def _as_decimal(val: Any) -> Decimal | None:
    if val is None or val == "":
        return None
    if isinstance(val, Decimal):
        return val
    try:
        d = Decimal(str(val))
        if d == 0:
            return None
        return d
    except (InvalidOperation, ValueError):
        return None


def _ensure_import_account(session: Session, profile: Profile) -> Account:
    existing = (
        session.query(Account)
        .filter(Account.profile_id == profile.id, Account.nickname == "Imported (Excel)")
        .first()
    )
    if existing:
        return existing
    acct = Account(
        profile_id=profile.id,
        kind="checking",
        nickname="Imported (Excel)",
        institution="Budget.xlsx",
        current_balance=Decimal("0"),
        is_cash_for_ifpp=False,
        include_in_net_worth=False,
    )
    session.add(acct)
    session.flush()
    return acct


def _category_by_code(session: Session, code: str, profile_id: int) -> Category | None:
    # system codes exact
    cat = session.query(Category).filter(Category.code == code).first()
    if cat:
        return cat
    # personal codes
    cat = (
        session.query(Category)
        .filter(Category.code == code, Category.profile_id == profile_id)
        .first()
    )
    return cat


def import_budget_xlsx(
    session: Session,
    path: Path | str,
    *,
    profile_slug: str = "personal",
    sheet_name: str = "Budget",
    since: date | None = None,
    dry_run: bool = False,
) -> ImportResult:
    path = Path(path)
    result = ImportResult()
    if not path.exists():
        result.errors.append(f"File not found: {path}")
        return result

    profile = session.query(Profile).filter(Profile.slug == profile_slug).first()
    if not profile:
        result.errors.append(f"Profile not found: {profile_slug}")
        return result

    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        result.errors.append(f"Cannot open workbook: {e}")
        return result

    if sheet_name not in wb.sheetnames:
        result.errors.append(f"Sheet '{sheet_name}' not found. Have: {wb.sheetnames[:10]}")
        wb.close()
        return result

    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    try:
        header_row1 = next(rows, None)
        header_row2 = next(rows, None)
    except StopIteration:
        result.errors.append("Sheet too short")
        wb.close()
        return result

    if not header_row2:
        result.errors.append("Missing header row 2")
        wb.close()
        return result

    # Map column index → category code
    col_map: dict[int, str] = {}
    for idx, cell in enumerate(header_row2):
        h = _norm_header(cell)
        if h in LEGACY_COLUMN_MAP and LEGACY_COLUMN_MAP[h]:
            col_map[idx] = LEGACY_COLUMN_MAP[h]
        elif h in ("amz", "apple", "target", "dicover", "discover", "amex", "hd", "wf"):
            # card movement columns — skip as expense categories (payment account noise)
            continue

    if not col_map:
        result.errors.append("No recognizable category columns in header row 2")
        wb.close()
        return result

    acct = None if dry_run else _ensure_import_account(session, profile)
    cat_cache: dict[str, Category | None] = {}

    def get_cat(code: str) -> Category | None:
        if code not in cat_cache:
            cat_cache[code] = _category_by_code(session, code, profile.id)
        return cat_cache[code]

    # existing external ids for idempotency
    existing_ids: set[str] = set()
    if not dry_run:
        for (ext,) in (
            session.query(Transaction.external_id)
            .filter(
                Transaction.profile_id == profile.id,
                Transaction.external_id.isnot(None),
            )
            .all()
        ):
            if ext:
                existing_ids.add(ext)

    for row in rows:
        result.rows_scanned += 1
        if not row:
            continue
        d = _as_date(row[0] if len(row) else None)
        if not d:
            continue
        if since and d < since:
            continue

        if result.date_from is None or d < result.date_from:
            result.date_from = d
        if result.date_to is None or d > result.date_to:
            result.date_to = d

        any_val = False
        for col_idx, code in col_map.items():
            if col_idx >= len(row):
                continue
            amt = _as_decimal(row[col_idx])
            if amt is None:
                continue
            any_val = True
            # Excel: expenses often negative already; income positive
            # If Food is stored as positive spend in some cells, normalize:
            # User sheet uses negative for outflows in most category cols.
            external_id = f"xlsx:{sheet_name}:{d.isoformat()}:{code}:{col_idx}:{amt}"
            if external_id in existing_ids:
                result.skipped_existing += 1
                continue

            cat = get_cat(code)
            if dry_run:
                result.transactions_created += 1
                continue

            assert acct is not None
            txn = Transaction(
                profile_id=profile.id,
                account_id=acct.id,
                category_id=cat.id if cat else None,
                txn_date=d,
                amount=amt,
                payee=f"Excel import ({code})",
                memo=f"Imported from {path.name} col={code}",
                status="cleared",
                external_id=external_id,
                is_transfer=code.startswith("SYS_TRANSFER"),
            )
            session.add(txn)
            existing_ids.add(external_id)
            result.transactions_created += 1

        if not any_val:
            result.skipped_empty += 1

    if not dry_run:
        session.flush()
        # Honesty: imported rows may match bills already paid — advance schedules (high confidence)
        if result.transactions_created > 0 and acct is not None:
            try:
                from financial_os.services.schedule_mark_paid import advance_schedules_after_import

                adv = advance_schedules_after_import(
                    session,
                    account_id=acct.id,
                    profile_id=profile.id,
                    auto_apply=True,
                )
                result.schedules_advanced = int(adv.get("advanced_count") or 0)
                result.schedules_advanced_names = [
                    str(x.get("name") or "") for x in (adv.get("advanced") or []) if x.get("name")
                ]
                result.schedule_advance_hint = adv.get("hint")
            except Exception as e:
                # Import still succeeded; surface advance failure without failing the import
                result.schedules_advanced = 0
                result.schedule_advance_error = str(e)[:300]
    wb.close()
    return result
